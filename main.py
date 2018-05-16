import os
import sqlite3 as sqlite
from openpyxl import Workbook
from openpyxl.styles import Font
import datetime
import tkinter.filedialog
import tkinter.messagebox
import gc

font = Font(color="00FF3300")
runningFile = ""


# 写入Excel头
def sqlite_to_workbook_with_head(i, cur, table, select_sql, workbook, errorPut):
    ws = workbook.worksheets[i]
    ws.title = table
    print("------- sheet create table -> " + table + " %s" % i)
    # enumerate针对一个可迭代对象，生成的是序号加上内容
    for colx, heading in enumerate(sqlite_get_col_names(cur, select_sql)):
        ws.cell(1, colx + 1, heading)
    for rowy, row in enumerate(query_by_sql(cur, select_sql)):
        for colx, text in enumerate(row):  # row是一行的内容
            try:
                ws.cell(rowy + 2, colx + 1, text)
            except:
                errorList = "file -> " + runningFile + "  :  sheet -> " + table + "\n"
                cell = ws.cell(rowy + 2, colx + 1, "Error Data")
                cell.font = font
                errorPut.write(errorList)
    del ws
    gc.collect()


# 写入Excel内容
def dump_db_to_excel(cur, workbook, errorPut):
    i = 0
    for tbl_name in [row[0] for row in query_by_sql(cur, "select tbl_name FROM sqlite_master where type = 'table'")]:
        select_sql = "select * from '%s'" % tbl_name
        if i > 0:
            workbook.create_sheet(tbl_name, i)
        sqlite_to_workbook_with_head(i, cur, tbl_name, select_sql, workbook, errorPut)
        i = i + 1


# 通过查询条件 查找 数据库 内容
def sqlite_get_col_names(cur, select_sql):
    cur.execute(select_sql)
    return [tuple[0] for tuple in cur.description]


# 查询数据库
def query_by_sql(cur, select_sql):
    cur.execute(select_sql)
    return cur.fetchall()


# 单个 转换 数据库 为 Excel 脚本
def main(dbpath, errorPut):
    xlspath = dbpath[:dbpath.rfind('.')] + '.xlsx'
    global runningFile
    runningFile = xlspath
    db = sqlite.connect(dbpath)
    cur = db.cursor()
    wb = Workbook()
    dump_db_to_excel(cur, wb, errorPut)
    print("end transfer <%s> --> <%s>" % (dbpath, xlspath))
    wb.save(xlspath)


# 程序入口 批量执行 path 一级目录下所有 .db 或者 .data 的数据库
if __name__ == '__main__':
    # 地址需要设置为目标存有数据库的文件夹
    global path
    path = tkinter.filedialog.askdirectory()
    errorPut = open(path + "/log.txt", 'w')
    startTime = datetime.datetime.now()
    errorPut.write("error list :\n")
    # path = "/Users/seapeak/Downloads/dblist"
    for fileName in os.listdir(path):
        if fileName.endswith('db') or fileName.endswith('data'):
            itemPath = os.path.join(path, fileName)
            print("start transfer -> " + itemPath)
            main(itemPath, errorPut)

    endTime = datetime.datetime.now()
    timeSpend = endTime - startTime
    print('sum time is -> %s' % timeSpend)
    errorPut.write('\n ------- transfer sum time = %s' % timeSpend + '\n')
    errorPut.close()
    tkinter.messagebox.showinfo("转换结束", "花费了 %s" % timeSpend + " 秒\nlog文件已经在文件夹中生成，可以查看转化失败的区域")
